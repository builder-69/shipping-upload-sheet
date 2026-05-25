"""
통합주문서 엑셀 파일 생성 모듈
"""
import pandas as pd
from datetime import datetime
from openpyxl.styles import numbers


def set_column_widths_and_formats(worksheet, df, column_width=20):
    """
    워크시트의 열 너비 설정 및 숫자 형식 지정
    
    Args:
        worksheet: openpyxl 워크시트 객체
        df: DataFrame
        column_width: 열 너비 (기본값 20)
    """
    # 숫자로 표시되어야 할 컬럼명 리스트
    number_columns = [
        '주문번호', '상품주문번호', '옵션상품번호', '핸드폰번호', '상품주문번호',
        '전화번호1', '수취인연락처1', '수취인연락처', '수취인 연락처', 
        '연락처', '핸드폰번호', '주문자 연락처', '우편번호'
    ]
    
    # 날짜로 표시되어야 할 컬럼명 리스트
    date_columns = [
        '결제일', '결제일자', '주문일시', '발주확인일', '발송기한', 
        '발송처리일', '송장출력일', '발송일', '배송예정일'
    ]
    
    # 모든 열의 너비를 설정
    for idx, col in enumerate(df.columns, 1):
        col_letter = worksheet.cell(row=1, column=idx).column_letter
        worksheet.column_dimensions[col_letter].width = column_width
        
        # 숫자로 표시되어야 하는 컬럼인 경우 텍스트 형식으로 설정
        if col in number_columns:
            for row_idx in range(2, len(df) + 2):  # 헤더 다음부터
                cell = worksheet.cell(row=row_idx, column=idx)
                # 값을 문자열로 변환하여 저장
                if cell.value is not None and cell.value != '':
                    cell.value = str(cell.value)
                    cell.number_format = '@'  # 텍스트 형식
        
        # 날짜 컬럼인 경우 원본 형식 유지
        elif col in date_columns:
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=idx)
                if cell.value is not None and cell.value != '':
                    # 이미 datetime 객체인 경우 형식 지정
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                    else:
                        # 텍스트로 저장된 경우 그대로 유지
                        cell.value = str(cell.value)
                        cell.number_format = '@'


def create_integrated_order_excel(
    naver_converted, naver_original,
    hyber_converted, hyber_original,
    ably_converted, ably_original,
    output_path=None
):
    """
    통합주문서 엑셀 파일 생성
    
    Args:
        naver_converted: 네이버 변환 DataFrame
        naver_original: 네이버 원본 DataFrame
        hyber_converted: 하이버 변환 DataFrame
        hyber_original: 하이버 원본 DataFrame
        ably_converted: 에이블리 변환 DataFrame
        ably_original: 에이블리 원본 DataFrame
        output_path: 출력 파일 경로 (None일 경우 자동 생성)
        
    Returns:
        str: 생성된 파일 경로
    """
    # 출력 파일명 생성 (날짜 포함)
    if output_path is None:
        today = datetime.now().strftime('%Y%m%d')
        output_path = f'/mnt/user-data/outputs/통합주문서_{today}.xlsx'
    
    # 시트1: 통합주문서 (네이버 + 하이버 + 에이블리 순서대로 연결)
    integrated_df = pd.concat([
        naver_converted,
        hyber_converted,
        ably_converted
    ], ignore_index=True)
    integrated_df = integrated_df.fillna('').replace(r'(?i)^\s*nan\s*$', '', regex=True)
    
    # 엑셀 파일 생성
    with pd.ExcelWriter(output_path, engine='openpyxl', date_format='YYYY-MM-DD HH:MM:SS', datetime_format='YYYY-MM-DD HH:MM:SS') as writer:
        # 시트1: 통합주문서
        integrated_df.to_excel(writer, sheet_name='통합주문서', index=False)
        ws_integrated = writer.sheets['통합주문서']
        set_column_widths_and_formats(ws_integrated, integrated_df, column_width=20)
        
        # 시트2: 네이버주문서 (원본)
        naver_original.to_excel(writer, sheet_name='네이버주문서', index=False)
        ws_naver = writer.sheets['네이버주문서']
        set_column_widths_and_formats(ws_naver, naver_original, column_width=20)
        
        # 시트3: 하이버주문서 (원본)
        hyber_original.to_excel(writer, sheet_name='하이버주문서', index=False)
        ws_hyber = writer.sheets['하이버주문서']
        set_column_widths_and_formats(ws_hyber, hyber_original, column_width=20)
        
        # 시트4: 에이블리주문서 (원본)
        ably_original.to_excel(writer, sheet_name='에이블리주문서', index=False)
        ws_ably = writer.sheets['에이블리주문서']
        set_column_widths_and_formats(ws_ably, ably_original, column_width=20)
    
    return output_path
